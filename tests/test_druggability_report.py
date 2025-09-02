import pandas as pd
from io import BytesIO
import xlsxwriter
from openpyxl import load_workbook


def test_pubmed_search(monkeypatch, dr_module):
    dr = dr_module
    monkeypatch.setattr(dr.Entrez, "esearch", lambda **k: "es")
    monkeypatch.setattr(dr.Entrez, "read", lambda h: {"Count": "1", "IdList": ["123456"]})
    monkeypatch.setattr(dr.Entrez, "efetch", lambda **k: "ef")
    records = [
        {
            "PMID": "123456",
            "Publication Type": ["Journal Article"],
            "Date of Publication": "2020 Jan 1",
            "MeSH Terms": ["Alzheimer Disease", "Amino Acids/*chemistry"],
            "Other Term": ["Dementia"],
            "Title": "Example",
            "Volume": "1",
            "Abstract": "Abs",
            "Affiliation": "Aff",
            "Author": "Auth",
            "Journal Title": "Journ",
            "Other Term Owner": "Owner",
            "Place of Publication": "Place",
            "Subset": "Sub",
            "Source": "Src",
            "Journal Title Abbreviation": "JTA",
            "Extra": "foo",
        }
    ]
    monkeypatch.setattr(dr.Medline, "parse", lambda h: records)
    df = dr.pubmed_search("GENE", "email@example.com")
    expected_cols = [
        "Abstract",
        "Affiliation",
        "Author",
        "Date of Publication",
        "Journal Title",
        "MeSH Terms",
        "Other Term",
        "Other Term Owner",
        "Place of Publication",
        "PMID",
        "Subset",
        "Source",
        "Journal Title Abbreviation",
        "Title",
        "Volume",
        "Journal Article",
        "Case Reports",
        "Clinical Trial",
        "Comparative Study",
        "Letter",
        "Meta-Analysis",
        "Review",
        "Year of Publication",
        "Neurodegeneration",
        "Major Keywords",
        "Chemistry",
    ]
    assert list(df.columns) == expected_cols
    row = df.iloc[0]
    assert row["Neurodegeneration"]
    assert row["Chemistry"]
    assert row["PMID"] == "https://www.ncbi.nlm.nih.gov/pubmed/123456/"


def test_write_excel_header(dr_module):
    dr = dr_module
    stream = BytesIO()
    wb = xlsxwriter.Workbook(stream, {"in_memory": True})
    ws = wb.add_worksheet()
    fmt = wb.add_format()
    headers = {"Single": (0, 0), "Merged": (1, 0, 1, 1)}
    dr.write_excel_header(headers, ws, fmt)
    wb.close()
    stream.seek(0)
    rb = load_workbook(stream)
    ws_r = rb.active
    assert ws_r["A1"].value == "Single"
    assert ws_r["A2"].value == "Merged"
    assert "A2:B2" in [str(r) for r in ws_r.merged_cells.ranges]


def test_get_single_excel(tmp_path, monkeypatch, dr_module):
    dr = dr_module
    dr.output_single_path = tmp_path
    dr.list_of_entries = pd.DataFrame(index=["P12345"])
    dr.pubmed_email = None
    dr.targetDB = "dummy.db"
    keys = {
        "ADME",
        "organ_expression",
        "dose_response",
        "domain_drugE",
        "bindingDB",
        "binding",
        "disease_exp",
        "isoforms_mod",
        "isoforms",
        "tissue_expression",
        "other",
        "mut",
        "kegg",
        "gwas",
        "domains",
        "tissue",
        "var",
        "open_target",
        "pdb",
        "pdb_blast",
        "phenotype",
        "commercials",
        "efficacy_bio",
        "disease",
        "reactome",
        "pockets",
        "alt_pockets",
        "percent_inhibition",
        "selectivity",
        "emax",
        "general_info",
    }
    empty = pd.DataFrame()

    def fake_features(uniprot_id, dbase=None):
        return {k: empty for k in keys}

    monkeypatch.setattr(dr.tf, "get_single_features", fake_features)
    target = pd.Series({"uniprot_ids": ["P12345"], "symbol": "GENE"})
    dr.get_single_excel(target)
    file_path = tmp_path / "GENE_P12345.xlsx"
    assert file_path.exists()
    wb = load_workbook(file_path)
    expected = [
        "General info",
        "Pubmed_search",
        "diseases",
        "open_target_association",
        "expression",
        "genotypes",
        "isoforms",
        "variants_mutants",
        "Structure",
    ]
    for sheet in expected:
        assert sheet in wb.sheetnames
